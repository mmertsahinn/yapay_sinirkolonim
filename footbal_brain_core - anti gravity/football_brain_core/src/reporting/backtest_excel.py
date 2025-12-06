"""
Backtest sonuçlarını Excel'e export et
PRD'ye uygun: Test sonuçları, doğruluk, hata analizleri
"""
from datetime import datetime, date
from typing import List, Dict, Optional, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, Reference
import logging
import os

from football_brain_core.src.db.connection import get_session
from football_brain_core.src.db.repositories import (
    MatchRepository, PredictionRepository, ResultRepository,
    ExplanationRepository, TeamRepository, LeagueRepository, MarketRepository
)
from football_brain_core.src.features.market_targets import MarketType
from football_brain_core.src.config import Config
from football_brain_core.src.inference.backtest import Backtester
from football_brain_core.src.models.multi_task_model import MultiTaskModel

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class BacktestExcelExporter:
    """
    Backtest sonuçlarını ve model performansını Excel'e export eder.
    PRD'ye uygun: Her maç için tahmin, gerçek, doğruluk, LLM açıklamaları.
    """
    
    def __init__(self, config: Optional[Config] = None):
        self.config = config or Config()
        os.makedirs(self.config.REPORTS_DIR, exist_ok=True)
    
    def export_backtest_results(
        self,
        backtest_results: Dict[str, Any],
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        model_version_id: Optional[int] = None
    ) -> str:
        """
        Backtest sonuçlarını Excel'e export et.
        
        Args:
            backtest_results: Backtester'dan gelen sonuçlar
            date_from: Başlangıç tarihi
            date_to: Bitiş tarihi
            model_version_id: Model versiyonu ID
        """
        session = get_session()
        try:
            rows = []
            
            # Her maç için detaylı satır oluştur
            for match_result in backtest_results.get("match_results", []):
                match_id = match_result.get("match_id")
                match = MatchRepository.get_by_id(session, match_id)
                
                if not match:
                    continue
                
                league = LeagueRepository.get_by_id(session, match.league_id)
                home_team = TeamRepository.get_by_id(session, match.home_team_id)
                away_team = TeamRepository.get_by_id(session, match.away_team_id)
                
                row = {
                    "Match ID": match.id,
                    "Date": match.match_date.strftime("%Y-%m-%d"),
                    "League": league.name if league else "",
                    "Home Team": home_team.name if home_team else "",
                    "Away Team": away_team.name if away_team else "",
                    "Home Score": match.home_score if match.home_score is not None else "",
                    "Away Score": match.away_score if match.away_score is not None else "",
                }
                
                predictions_dict = match_result.get("predictions", {})
                actuals_dict = match_result.get("actuals", {})
                correct_dict = match_result.get("correct", {})
                
                # Her market için kolonlar ekle
                for market_name, pred in predictions_dict.items():
                    row[f"{market_name} - Prediction"] = pred
                    row[f"{market_name} - Actual"] = actuals_dict.get(market_name, "")
                    is_correct = correct_dict.get(market_name, False)
                    row[f"{market_name} - Correct"] = "Yes" if is_correct else "No"
                
                # LLM açıklamaları
                explanations = ExplanationRepository.get_by_match(session, match.id)
                for explanation in explanations:
                    market = MarketRepository.get_by_id(session, explanation.market_id)
                    if market:
                        row[f"{market.name} - Explanation"] = explanation.llm_output
                        
                        # Model bilgileri
                        if explanation.summary_stats:
                            stats = explanation.summary_stats
                            row[f"{market.name} - Model"] = stats.get("best_model", "N/A").upper()
                            row[f"{market.name} - GPT Time"] = stats.get("gpt_time", "N/A")
                            row[f"{market.name} - Grok Time"] = stats.get("grok_time", "N/A")
                
                rows.append(row)
            
            # Özet sayfası için metrikler
            summary_data = {
                "Metric": [],
                "Value": []
            }
            
            accuracy_by_market = backtest_results.get("accuracy_by_market", {})
            for market, accuracy in accuracy_by_market.items():
                summary_data["Metric"].append(f"{market} - Accuracy")
                summary_data["Value"].append(f"{accuracy:.2%}")
            
            summary_data["Metric"].append("Total Matches")
            summary_data["Value"].append(backtest_results.get("total_matches", 0))
            
            # DataFrame oluştur
            df = pd.DataFrame(rows)
            df_summary = pd.DataFrame(summary_data)
            
            # Excel'e yaz
            if not date_from:
                date_from = date.today()
            if not date_to:
                date_to = date_from
            
            output_path = os.path.join(
                self.config.REPORTS_DIR,
                f"backtest_results_{date_from}_{date_to}.xlsx"
            )
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Test Results', index=False)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Formatla
            self._format_backtest_excel(output_path)
            
            logger.info(f"Backtest sonuçları export edildi: {output_path}")
            return output_path
        
        finally:
            session.close()
    
    def export_model_performance(
        self,
        evaluation_results: Dict[str, Dict[str, float]],
        model_version: str = "v1.0"
    ) -> str:
        """
        Model performans metriklerini Excel'e export et.
        """
        rows = []
        
        for market_name, metrics in evaluation_results.items():
            row = {
                "Market": market_name,
                "Accuracy": metrics.get("accuracy", 0.0),
                "Brier Score": metrics.get("brier_score", 0.0),
                "Log Loss": metrics.get("log_loss", 0.0),
                "Num Samples": metrics.get("num_samples", 0)
            }
            rows.append(row)
        
        df = pd.DataFrame(rows)
        
        output_path = os.path.join(
            self.config.REPORTS_DIR,
            f"model_performance_{model_version}.xlsx"
        )
        
        df.to_excel(output_path, index=False)
        self._format_performance_excel(output_path)
        
        logger.info(f"Model performans metrikleri export edildi: {output_path}")
        return output_path
    
    def _format_backtest_excel(self, file_path: str):
        """Backtest Excel'ini formatla"""
        wb = load_workbook(file_path)
        
        # Test Results sayfası
        if 'Test Results' in wb.sheetnames:
            ws = wb['Test Results']
            
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Header formatla
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
            
            # Doğruluk kolonlarını formatla
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if "Correct" in str(cell.column_letter) or "Correct" in str(cell.value):
                        if cell.value == "Yes":
                            cell.fill = green_fill
                        elif cell.value == "No":
                            cell.fill = red_fill
        
        # Summary sayfası
        if 'Summary' in wb.sheetnames:
            ws = wb['Summary']
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
        
        wb.save(file_path)
    
    def _format_performance_excel(self, file_path: str):
        """Performans Excel'ini formatla"""
        wb = load_workbook(file_path)
        ws = wb.active
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
        
        wb.save(file_path)







