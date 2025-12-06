"""
Takım Analizi Excel Export
- Her takımın pattern'leri
- Takım çiftlerinin ilişkileri
- Her iterasyondan sonra güncellenmiş analizler
"""
from typing import Dict, List, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import logging
import os

from football_brain_core.src.config import Config

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class TeamAnalysisExcelExporter:
    """
    Takım analizlerini Excel'e export eder.
    Her iterasyondan sonra güncellenmiş takım pattern'leri ve ilişkileri.
    """
    
    def __init__(self, config: Config = None):
        self.config = config or Config()
        os.makedirs(self.config.REPORTS_DIR, exist_ok=True)
    
    def export_team_patterns(
        self,
        team_patterns: Dict[int, Dict[str, Any]],
        iteration: int,
        season: int
    ) -> str:
        """
        Takım pattern'lerini Excel'e export et.
        """
        rows = []
        
        for team_id, team_data in team_patterns.items():
            team_name = team_data.get("team_name", f"Team {team_id}")
            patterns = team_data.get("patterns", {})
            
            for market_name, pattern_data in patterns.items():
                row = {
                    "Iteration": iteration,
                    "Season": season,
                    "Team ID": team_id,
                    "Team Name": team_name,
                    "Market": market_name,
                }
                
                # Ev sahibi pattern'leri
                home_patterns = pattern_data.get("home_patterns", {})
                for outcome, prob in home_patterns.items():
                    row[f"Home - {outcome}"] = f"{prob:.1%}"
                
                # Deplasman pattern'leri
                away_patterns = pattern_data.get("away_patterns", {})
                for outcome, prob in away_patterns.items():
                    row[f"Away - {outcome}"] = f"{prob:.1%}"
                
                # Güçlü yönler
                strengths = pattern_data.get("strengths", [])
                row["Strengths"] = "; ".join(strengths)
                
                # Zayıf yönler
                weaknesses = pattern_data.get("weaknesses", [])
                row["Weaknesses"] = "; ".join(weaknesses)
                
                rows.append(row)
        
        df = pd.DataFrame(rows)
        
        output_path = os.path.join(
            self.config.REPORTS_DIR,
            f"team_patterns_iteration_{iteration}_season_{season}.xlsx"
        )
        
        df.to_excel(output_path, index=False)
        self._format_team_excel(output_path)
        
        logger.info(f"Takım pattern'leri export edildi: {output_path}")
        return output_path
    
    def export_team_relationships(
        self,
        relationships: List[Dict[str, Any]],
        iteration: int,
        season: int
    ) -> str:
        """
        Takım ilişkilerini Excel'e export et.
        """
        rows = []
        
        for rel in relationships:
            row = {
                "Iteration": iteration,
                "Season": season,
                "Team A": rel.get("team_a_name", ""),
                "Team B": rel.get("team_b_name", ""),
                "Market": rel.get("market", ""),
                "Total Matches": rel.get("total_matches", 0),
                "Relationship Type": rel.get("relationship_type", ""),
                "Dominance": rel.get("dominance", ""),
            }
            
            # Team A ev sahibi pattern'leri
            a_home = rel.get("team_a_as_home", {})
            for outcome, count in a_home.items():
                row[f"Team A Home - {outcome}"] = count
            
            # Team A deplasman pattern'leri
            a_away = rel.get("team_a_as_away", {})
            for outcome, count in a_away.items():
                row[f"Team A Away - {outcome}"] = count
            
            rows.append(row)
        
        df = pd.DataFrame(rows)
        
        output_path = os.path.join(
            self.config.REPORTS_DIR,
            f"team_relationships_iteration_{iteration}_season_{season}.xlsx"
        )
        
        df.to_excel(output_path, index=False)
        self._format_relationship_excel(output_path)
        
        logger.info(f"Takım ilişkileri export edildi: {output_path}")
        return output_path
    
    def export_comprehensive_analysis(
        self,
        team_patterns: Dict[int, Dict[str, Any]],
        relationships: List[Dict[str, Any]],
        iteration: int,
        season: int,
        error_summary: Dict[str, Any]
    ) -> str:
        """
        Kapsamlı analiz Excel'i: Pattern'ler + İlişkiler + Hata analizi
        """
        output_path = os.path.join(
            self.config.REPORTS_DIR,
            f"comprehensive_analysis_iteration_{iteration}_season_{season}.xlsx"
        )
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 1. Takım Pattern'leri
            pattern_rows = []
            for team_id, team_data in team_patterns.items():
                for market_name, pattern_data in team_data.get("patterns", {}).items():
                    pattern_rows.append({
                        "Team": team_data.get("team_name", ""),
                        "Market": market_name,
                        "Home Patterns": str(pattern_data.get("home_patterns", {})),
                        "Away Patterns": str(pattern_data.get("away_patterns", {})),
                        "Strengths": "; ".join(pattern_data.get("strengths", [])),
                        "Weaknesses": "; ".join(pattern_data.get("weaknesses", []))
                    })
            
            if pattern_rows:
                pd.DataFrame(pattern_rows).to_excel(writer, sheet_name='Team Patterns', index=False)
            
            # 2. Takım İlişkileri
            if relationships:
                rel_rows = []
                for rel in relationships:
                    rel_rows.append({
                        "Team A": rel.get("team_a_name", ""),
                        "Team B": rel.get("team_b_name", ""),
                        "Market": rel.get("market", ""),
                        "Matches": rel.get("total_matches", 0),
                        "Type": rel.get("relationship_type", ""),
                        "Dominance": rel.get("dominance", "")
                    })
                pd.DataFrame(rel_rows).to_excel(writer, sheet_name='Team Relationships', index=False)
            
            # 3. Hata Özeti
            if error_summary:
                error_rows = []
                error_rows.append({"Metric": "Total Errors", "Value": error_summary.get("total_errors", 0)})
                error_rows.append({"Metric": "Bias Count", "Value": error_summary.get("bias_count", 0)})
                error_rows.append({"Metric": "Variance Count", "Value": error_summary.get("variance_count", 0)})
                
                for error_type, count in error_summary.get("error_types", {}).items():
                    error_rows.append({"Metric": f"Error Type: {error_type}", "Value": count})
                
                for suggestion in error_summary.get("suggested_improvements", []):
                    error_rows.append({"Metric": "Suggestion", "Value": suggestion})
                
                pd.DataFrame(error_rows).to_excel(writer, sheet_name='Error Analysis', index=False)
        
        self._format_comprehensive_excel(output_path)
        
        logger.info(f"Kapsamlı analiz export edildi: {output_path}")
        return output_path
    
    def _format_team_excel(self, file_path: str):
        """Takım Excel'ini formatla"""
        wb = load_workbook(file_path)
        ws = wb.active
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
        
        wb.save(file_path)
    
    def _format_relationship_excel(self, file_path: str):
        """İlişki Excel'ini formatla"""
        self._format_team_excel(file_path)
    
    def _format_comprehensive_excel(self, file_path: str):
        """Kapsamlı Excel'i formatla"""
        wb = load_workbook(file_path)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
        
        wb.save(file_path)


