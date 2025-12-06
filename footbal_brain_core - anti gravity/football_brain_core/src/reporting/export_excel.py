from datetime import datetime, date
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import logging
import os

from football_brain_core.src.db.connection import get_session
from football_brain_core.src.db.repositories import (
    MatchRepository, PredictionRepository, ResultRepository,
    ExplanationRepository, TeamRepository, LeagueRepository, MarketRepository
)
from football_brain_core.src.features.market_targets import MarketType
from football_brain_core.src.config import Config

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelExporter:
    def __init__(self, config: Optional[Config] = None):
        self.config = config or Config()
        os.makedirs(self.config.REPORTS_DIR, exist_ok=True)
    
    def export_predictions(
        self,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        league_ids: Optional[List[int]] = None,
        model_version_id: Optional[int] = None
    ) -> str:
        session = get_session()
        try:
            if not date_from:
                date_from = date.today()
            if not date_to:
                date_to = date_from
            
            matches = MatchRepository.get_by_date_range(session, date_from, date_to)
            
            if league_ids:
                matches = [m for m in matches if m.league_id in league_ids]
            
            rows = []
            
            for match in matches:
                league = LeagueRepository.get_by_id(session, match.league_id)
                home_team = TeamRepository.get_by_id(session, match.home_team_id)
                away_team = TeamRepository.get_by_id(session, match.away_team_id)
                
                row = {
                    "Match ID": match.id,
                    "Date": match.match_date.strftime("%Y-%m-%d"),
                    "League": league.name if league else "",
                    "Home Team": home_team.name if home_team else "",
                    "Away Team": away_team.name if away_team else "",
                    "Home Score": match.home_score if match.home_score is not None else "",
                    "Away Score": match.away_score if match.away_score is not None else "",
                }
                
                predictions = PredictionRepository.get_by_match(session, match.id)
                if model_version_id:
                    predictions = [p for p in predictions if p.model_version_id == model_version_id]
                
                results = ResultRepository.get_by_match(session, match.id)
                explanations = ExplanationRepository.get_by_match(session, match.id)
                
                for pred in predictions:
                    market = MarketRepository.get_by_id(session, pred.market_id)
                    if not market:
                        continue
                    
                    market_name = market.name
                    row[f"{market_name} - Prediction"] = pred.predicted_outcome
                    row[f"{market_name} - Probability"] = pred.p_hat if pred.p_hat else ""
                    
                    result = next((r for r in results if r.market_id == market.id), None)
                    if result:
                        row[f"{market_name} - Actual"] = result.actual_outcome
                        is_correct = pred.predicted_outcome == result.actual_outcome
                        row[f"{market_name} - Correct"] = "Yes" if is_correct else "No"
                    else:
                        row[f"{market_name} - Actual"] = ""
                        row[f"{market_name} - Correct"] = ""
                    
                    explanation = next(
                        (e for e in explanations if e.market_id == market.id), None
                    )
                    if explanation:
                        row[f"{market_name} - Explanation"] = explanation.llm_output
                        
                        # Model bilgisini ayrı kolonlara ekle
                        if explanation.summary_stats:
                            stats = explanation.summary_stats
                            row[f"{market_name} - Model (GPT/Grok)"] = stats.get("best_model", "N/A").upper()
                            row[f"{market_name} - GPT Time (s)"] = stats.get("gpt_time", "N/A")
                            row[f"{market_name} - Grok Time (s)"] = stats.get("grok_time", "N/A")
                            row[f"{market_name} - Best Time (s)"] = stats.get("best_time", "N/A")
                            
                            # Her iki modelin açıklamasını da göster
                            if stats.get("gpt_explanation"):
                                row[f"{market_name} - GPT Explanation"] = stats.get("gpt_explanation", "")
                            if stats.get("grok_explanation"):
                                row[f"{market_name} - Grok Explanation"] = stats.get("grok_explanation", "")
                        else:
                            row[f"{market_name} - Model (GPT/Grok)"] = "N/A"
                    else:
                        row[f"{market_name} - Explanation"] = ""
                        row[f"{market_name} - Model (GPT/Grok)"] = ""
                
                rows.append(row)
            
            df = pd.DataFrame(rows)
            
            output_path = os.path.join(
                self.config.REPORTS_DIR,
                f"predictions_{date_from}_{date_to}.xlsx"
            )
            
            # Excel'e yaz (birden fazla sayfa)
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Predictions', index=False)
                
                # Özet sayfası ekle
                summary_data = self._create_summary_sheet(session, matches, predictions, results)
                if summary_data:
                    df_summary = pd.DataFrame(summary_data)
                    df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            self._format_excel(output_path)
            
            logger.info(f"Exported predictions to {output_path}")
            return output_path
        finally:
            session.close()
    
    def _create_summary_sheet(
        self,
        session,
        matches: List,
        predictions: List,
        results: List
    ) -> Optional[Dict[str, List]]:
        """Özet sayfası için veri oluştur"""
        if not predictions:
            return None
        
        summary = {
            "Metric": [],
            "Value": []
        }
        
        # Toplam maç sayısı
        summary["Metric"].append("Total Matches")
        summary["Value"].append(len(matches))
        
        # Market bazlı doğruluk
        market_correct = {}
        market_total = {}
        
        for pred in predictions:
            market = MarketRepository.get_by_id(session, pred.market_id)
            if not market:
                continue
            
            market_name = market.name
            if market_name not in market_total:
                market_total[market_name] = 0
                market_correct[market_name] = 0
            
            market_total[market_name] += 1
            
            # Gerçek sonuçla karşılaştır
            result = next((r for r in results if r.market_id == market.id and r.match_id == pred.match_id), None)
            if result and pred.predicted_outcome == result.actual_outcome:
                market_correct[market_name] += 1
        
        for market_name in market_total:
            accuracy = market_correct[market_name] / market_total[market_name] if market_total[market_name] > 0 else 0
            summary["Metric"].append(f"{market_name} - Accuracy")
            summary["Value"].append(f"{accuracy:.2%}")
            summary["Metric"].append(f"{market_name} - Correct/Total")
            summary["Value"].append(f"{market_correct[market_name]}/{market_total[market_name]}")
        
        return summary
    
    def _format_excel(self, file_path: str):
        wb = load_workbook(file_path)
        ws = wb.active
        
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if "Correct" in str(cell.value) and cell.value == "Yes":
                    cell.fill = green_fill
                elif "Correct" in str(cell.value) and cell.value == "No":
                    cell.fill = red_fill
        
        wb.save(file_path)

