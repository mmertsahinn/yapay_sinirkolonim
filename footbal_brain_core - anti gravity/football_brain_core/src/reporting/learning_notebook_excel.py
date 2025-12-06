"""
PRD: Excel Öğrenme Defteri
Her maç + market için:
- Lig, tarih, takımlar, skor
- Market tipi ve seçilen outcome etiketi
- Doğru/yanlış bilgisi (renk/flags)
- Basit form özetleri (son 5 maç puan ort., gol farkı vb.)
- LLM senaryosu (kısa metin)
"""
from datetime import date, datetime
from typing import List, Dict, Optional, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
import logging

from src.db.connection import get_session
from src.db.repositories import (
    MatchRepository, PredictionRepository, ResultRepository,
    ExplanationRepository, TeamRepository, LeagueRepository,
    MarketRepository
)
from src.features.feature_builder import FeatureBuilder

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class LearningNotebookExporter:
    """
    PRD: Excel Öğrenme Defteri
    Modelin öğrenme sürecini takip etmek için detaylı Excel raporu
    """
    
    def __init__(self, output_dir: Optional[str] = None):
        self.output_dir = Path(output_dir) if output_dir else Path("reports")
        self.output_dir.mkdir(exist_ok=True)
        self.feature_builder = FeatureBuilder()
    
    def export_learning_notebook(
        self,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        league_ids: Optional[List[int]] = None
    ) -> str:
        """
        PRD formatında Excel Öğrenme Defteri oluştur
        """
        session = get_session()
        try:
            # Tarih aralığı
            if not date_from:
                date_from = date.today() - pd.Timedelta(days=7)
            if not date_to:
                date_to = date.today()
            
            # Maçları al
            matches = MatchRepository.get_by_date_range(session, date_from, date_to)
            if league_ids:
                matches = [m for m in matches if m.league_id in league_ids]
            
            # Her maç için detaylı satırlar oluştur
            rows = []
            
            for match in matches:
                league = LeagueRepository.get_by_id(session, match.league_id)
                home_team = TeamRepository.get_by_id(session, match.home_team_id)
                away_team = TeamRepository.get_by_id(session, match.away_team_id)
                
                # Form özetleri
                home_form = self._get_form_summary(match.home_team_id, match.match_date, session)
                away_form = self._get_form_summary(match.away_team_id, match.match_date, session)
                
                # Tahminler ve sonuçlar
                predictions = PredictionRepository.get_by_match(session, match.id)
                results = ResultRepository.get_by_match(session, match.id)
                explanations = ExplanationRepository.get_by_match(session, match.id)
                
                # Her market için ayrı satır
                for pred in predictions:
                    market = MarketRepository.get_by_id(session, pred.market_id)
                    if not market:
                        continue
                    
                    result = next((r for r in results if r.market_id == pred.market_id), None)
                    explanation = next((e for e in explanations if e.market_id == pred.market_id), None)
                    
                    # Doğru/yanlış
                    is_correct = None
                    if result:
                        is_correct = pred.predicted_outcome == result.actual_outcome
                    
                    # Satır oluştur
                    row = {
                        "Lig": league.name if league else "",
                        "Tarih": match.match_date.strftime("%Y-%m-%d"),
                        "Ev Takımı": home_team.name if home_team else "",
                        "Deplasman Takımı": away_team.name if away_team else "",
                        "Ev Skor": match.home_score if match.home_score is not None else "",
                        "Deplasman Skor": match.away_score if match.away_score is not None else "",
                        "Market": market.name,
                        "Tahmin": pred.predicted_outcome,
                        "Gerçek Sonuç": result.actual_outcome if result else "",
                        "Doğru/Yanlış": "✅ Doğru" if is_correct else ("❌ Yanlış" if is_correct is False else "⏳ Beklemede"),
                        # Form özetleri
                        "Ev Takımı - Son 5 Maç Puan Ort.": home_form.get("avg_points", ""),
                        "Ev Takımı - Gol Farkı": home_form.get("goal_diff", ""),
                        "Deplasman Takımı - Son 5 Maç Puan Ort.": away_form.get("avg_points", ""),
                        "Deplasman Takımı - Gol Farkı": away_form.get("goal_diff", ""),
                        # LLM Senaryosu
                        "LLM Senaryosu": explanation.llm_output[:200] + "..." if explanation and explanation.llm_output else "",
                    }
                    
                    rows.append(row)
            
            # DataFrame oluştur
            df = pd.DataFrame(rows)
            
            # Excel'e yaz
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.output_dir / f"ogrenme_defteri_{timestamp}.xlsx"
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Öğrenme Defteri', index=False)
            
            # Formatla
            self._format_learning_notebook(output_path)
            
            logger.info(f"✅ Öğrenme Defteri oluşturuldu: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Öğrenme Defteri oluşturma hatası: {e}")
            import traceback
            traceback.print_exc()
            raise
        finally:
            session.close()
    
    def _get_form_summary(
        self,
        team_id: int,
        match_date: datetime,
        session
    ) -> Dict[str, Any]:
        """Takımın son 5 maç form özeti"""
        try:
            features = self.feature_builder.build_team_features(team_id, match_date, session)
            
            return {
                "avg_points": round(features.get("avg_points", 0), 2),
                "goal_diff": round(features.get("avg_goal_diff", 0), 2),
                "win_rate": round(features.get("win_rate", 0), 2),
                "avg_goals_scored": round(features.get("avg_goals_scored", 0), 2),
                "avg_goals_conceded": round(features.get("avg_goals_conceded", 0), 2),
            }
        except:
            return {
                "avg_points": "",
                "goal_diff": "",
                "win_rate": "",
                "avg_goals_scored": "",
                "avg_goals_conceded": "",
            }
    
    def _format_learning_notebook(self, file_path: Path):
        """Excel dosyasını formatla (renkler, fontlar)"""
        try:
            wb = load_workbook(file_path)
            ws = wb['Öğrenme Defteri']
            
            # Başlık satırını kalın yap
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Doğru/Yanlış kolonunu renklendir
            correct_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == "Doğru/Yanlış":
                    correct_col = idx
                    break
            
            if correct_col:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=correct_col, max_col=correct_col):
                    for cell in row:
                        if "✅" in str(cell.value):
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif "❌" in str(cell.value):
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif "⏳" in str(cell.value):
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            # Kolon genişliklerini ayarla
            column_widths = {
                "A": 15,  # Lig
                "B": 12,  # Tarih
                "C": 20,  # Ev Takımı
                "D": 20,  # Deplasman Takımı
                "E": 10,  # Ev Skor
                "F": 10,  # Deplasman Skor
                "G": 20,  # Market
                "H": 15,  # Tahmin
                "I": 15,  # Gerçek Sonuç
                "J": 15,  # Doğru/Yanlış
                "K": 20,  # Ev Form
                "L": 15,  # Ev Gol Farkı
                "M": 20,  # Deplasman Form
                "N": 15,  # Deplasman Gol Farkı
                "O": 50,  # LLM Senaryosu
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Satır yüksekliği
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                ws.row_dimensions[row[0].row].height = 20
            
            wb.save(file_path)
            
        except Exception as e:
            logger.error(f"Excel formatlama hatası: {e}")






