"""
📝 EVRİM LOGGER - Detaylı LoRA Günlüğü
======================================

Her şeyi kaydeder:
- LoRA doğumları (anne, baba, genetik bilgi)
- LoRA ölümleri (sebep, yaş, fitness)
- Çiftleşmeler (kimle, nasıl)
- Mutasyonlar
- Spontane doğumlar (alien!)
- Şanslı kurtuluşlar
- Popülasyon istatistikleri (maç maç)
"""

import json
import os
from datetime import datetime
from typing import Dict, List, Optional
import pandas as pd


class EvolutionLogger:
    """
    Tüm evrim olaylarını detaylıca kaydeder
    """
    
    def __init__(self, log_dir: str = "evolution_logs"):
        self.log_dir = log_dir
        os.makedirs(log_dir, exist_ok=True)
        
        # Dosya yolları - SABİT İSİMLER! (üzerine yazar)
        self.main_log_file = os.path.join(log_dir, "evolution_log.txt")
        self.json_log_file = os.path.join(log_dir, "evolution_data.json")
        self.population_csv = os.path.join(log_dir, "population_history.csv")  # Excel olacak!
        self.events_csv = os.path.join(log_dir, "evolution_events.csv")  # Excel olacak!
        
        # Hafıza
        self.all_events = []
        self.population_history = []
        self.detailed_lora_history = []  # 🏷️ Her LoRA her maçta bir satır!
        self.resurrection_eras = []  # ⚡ Diriltme dönemleri (Excel ayraç için!)
        self.match_count = 0
        
        # 💀 ÖLÜM RAPORU (CANLI EXCEL!)
        self.death_report_file = os.path.join(log_dir, "OLUM_RAPORU_CANLI.xlsx")
        self._init_death_report_excel()
        
        # İlk log
        self._write_header()
    
    def _write_header(self):
        """Log dosyasının başlığı"""
        with open(self.main_log_file, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("🧬 KAOTİK EVRİMSEL LoRA SİSTEMİ - EVRIM GÜNLÜĞÜ\n")
            f.write("=" * 80 + "\n")
            f.write(f"Başlangıç Zamanı: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 80 + "\n\n")
    
    def log_birth(self, child_lora, parent1_lora=None, parent2_lora=None, 
                  mutation_applied=False, birth_type="crossover"):
        """
        LoRA doğumu logla
        
        birth_type: "crossover", "spontaneous", "emergency"
        """
        event = {
            'match': self.match_count,
            'timestamp': datetime.now().isoformat(),
            'type': 'birth',
            'birth_type': birth_type,
            'child_id': child_lora.id,
            'child_name': child_lora.name,
            'child_generation': child_lora.generation,
            'mutation_applied': mutation_applied
        }
        
        # Log mesajı
        if birth_type == "crossover":
            event['parent1_id'] = parent1_lora.id
            event['parent1_name'] = parent1_lora.name
            event['parent1_fitness'] = parent1_lora.get_recent_fitness()
            event['parent2_id'] = parent2_lora.id
            event['parent2_name'] = parent2_lora.name
            event['parent2_fitness'] = parent2_lora.get_recent_fitness()
            
            msg = f"\n{'='*80}\n"
            msg += f"🐣 YENİ LoRA DOĞDU! (Maç #{self.match_count})\n"
            msg += f"{'='*80}\n"
            msg += f"  Çocuk:\n"
            msg += f"    • İsim: {child_lora.name}\n"
            msg += f"    • ID: {child_lora.id}\n"
            msg += f"    • Generasyon: {child_lora.generation}\n"
            msg += f"\n"
            msg += f"  Anne:\n"
            msg += f"    • İsim: {parent1_lora.name}\n"
            msg += f"    • Fitness: {parent1_lora.get_recent_fitness():.3f}\n"
            msg += f"    • Generasyon: {parent1_lora.generation}\n"
            spec1 = getattr(parent1_lora, 'specialization', None)
            if spec1:
                msg += f"    • Uzmanlık: {spec1}\n"
            msg += f"\n"
            msg += f"  Baba:\n"
            msg += f"    • İsim: {parent2_lora.name}\n"
            msg += f"    • Fitness: {parent2_lora.get_recent_fitness():.3f}\n"
            msg += f"    • Generasyon: {parent2_lora.generation}\n"
            spec2 = getattr(parent2_lora, 'specialization', None)
            if spec2:
                msg += f"    • Uzmanlık: {spec2}\n"
            msg += f"\n"
            
            # Mutasyon detayı
            if mutation_applied:
                msg += f"  🧬 MUTASYON: ✅ EVET\n"
                msg += f"    • Bazı genler rastgele değiştirildi\n"
                msg += f"    • Beklenmedik yetenekler ortaya çıkabilir!\n"
            else:
                msg += f"  🧬 MUTASYON: ❌ HAYIR (saf çiftleşme)\n"
            
            msg += f"\n  💡 BEKLENTİ:\n"
            if spec1 and spec2 and spec1 != spec2:
                msg += f"    • {spec1} + {spec2} → Hibrit uzman olabilir!\n"
            elif spec1 and spec2 and spec1 == spec2:
                msg += f"    • İki {spec1} → Süper {spec1} olabilir!\n"
            else:
                msg += f"    • Ebeveynlerin en iyi genlerini taşıyabilir\n"
            
            msg += f"{'='*80}\n"
        
        elif birth_type == "spontaneous":
            msg = f"\n{'='*80}\n"
            msg += f"👽 ALIEN LoRA HİÇLİKTEN DOĞDU! (Maç #{self.match_count})\n"
            msg += f"{'='*80}\n"
            msg += f"  • İsim: {child_lora.name}\n"
            msg += f"  • ID: {child_lora.id}\n"
            msg += f"  • Ebeveyn: YOK (spontane doğum)\n"
            msg += f"  • Bu LoRA tamamen rastgele parametrelerle başladı!\n"
            msg += f"{'='*80}\n"
        
        elif birth_type == "emergency":
            msg = f"\n{'='*80}\n"
            msg += f"🚨 ACİL DOĞUM! (Maç #{self.match_count})\n"
            msg += f"{'='*80}\n"
            msg += f"  • İsim: {child_lora.name}\n"
            msg += f"  • ID: {child_lora.id}\n"
            msg += f"  • Sebep: Popülasyon çok azaldı (< minimum)\n"
            msg += f"{'='*80}\n"
        
        self._write_log(msg)
        self.all_events.append(event)
    
    def log_death(self, lora, reason="low_fitness", lucky_survived=False, death_reason_detail=None, physics_data=None):
        """
        LoRA ölümü logla
        
        reason: "low_fitness", "overpopulation", "forced"
        death_reason_detail: Detaylı ölüm sebebi
        physics_data: (Opsiyonel) Dışarıdan gelen fizik verileri dictionary'si
        """
        # Dirilme geçmişi varsa etiketle
        resurrection_tag = ""
        if hasattr(lora, 'resurrection_count') and lora.resurrection_count > 0:
            resurrection_tag = f" [⚡ {lora.resurrection_count}. Ölüm - Daha önce {lora.resurrection_count}x dirildi]"
        
        # 🔬 TES SKORLARI HESAPLA!
        from lora_system.tes_scoreboard import tes_scoreboard
        try:
            tes_data = tes_scoreboard.calculate_tes_score(lora, [], collective_memory=None)
        except:
            tes_data = {'total_tes': 0.0, 'darwin': 0.0, 'einstein': 0.0, 'newton': 0.0, 'lora_type': 'DENGELI⚖️'}
        
        # 🎭 FİZİK ARKETİPİ (Eski sistem - fluid temperament bazlı)
        from lora_system.physics_based_archetypes import physics_archetypes
        physics_arch = physics_archetypes.determine_archetype_from_physics(lora)
        
        # 🌊 PARÇACIK ARKETİPİ (YENİ! - parçacık fiziği bazlı)
        from lora_system.particle_archetypes import particle_archetypes
        particle_arch_data = particle_archetypes.get_archetype_from_lora(lora)
        particle_arch = particle_arch_data['primary_archetype']
        
        # ⚡ LIFE ENERGY
        life_energy = getattr(lora, 'life_energy', 0.0)
        
        # 🌊 PARÇACIK FİZİĞİ VERİLERİ (Öncelik: physics_data > lora attributes > default)
        if physics_data:
             langevin_temp = physics_data.get('langevin_temp', 0.01)
             nose_hoover_xi = physics_data.get('nose_hoover_xi', 0.0)
             kinetic_energy = physics_data.get('kinetic_energy', 0.0)
             om_action = physics_data.get('om_action', 0.0)
             lazarus_lambda = physics_data.get('lazarus_lambda', 0.5)
             ghost_potential = physics_data.get('ghost_potential', 0.0)
        else:
             langevin_temp = getattr(lora, '_langevin_temp', 0.01)
             nose_hoover_xi = getattr(lora, '_nose_hoover_xi', 0.0)
             kinetic_energy = getattr(lora, '_kinetic_energy', 0.0)
             om_action = getattr(lora, '_om_action', 0.0)
             lazarus_lambda = getattr(lora, '_lazarus_lambda', 0.5)
             ghost_potential = getattr(lora, '_ghost_potential', 0.0)
        
        event = {
            'match': self.match_count,
            'timestamp': datetime.now().isoformat(),
            'type': 'death',
            'lora_id': lora.id,
            'lora_name': lora.name,
            'age_in_matches': self.match_count - lora.birth_match,
            'age_matches': (self.match_count - lora.birth_match) if self.match_count > lora.birth_match else 0,
            'final_fitness': lora.get_recent_fitness(),
            'generation': lora.generation,
            'tes_scores': tes_data,  # 🔬 TES!
            'life_energy': life_energy,  # ⚡ Energy!
            'langevin_temp': langevin_temp,     # 🌊 Langevin T
            'nose_hoover_xi': nose_hoover_xi,   # 🌊 Nosé-Hoover ξ
            'kinetic_energy': kinetic_energy,   # 🌊 KE
            'om_action': om_action,             # 🌀 Onsager-Machlup
            'lazarus_lambda': lazarus_lambda,   # 🧟 Lazarus Λ
            'ghost_potential': ghost_potential, # 👻 Ghost U
            'physics_archetype': physics_arch,  # 🎭 Fizik!
            'particle_archetype': particle_arch, # 🌊 Parçacık!
            'reason': reason,
            'death_detail': death_reason_detail or reason,
            'lucky_survived': lucky_survived,
            'resurrection_count': getattr(lora, 'resurrection_count', 0),
            'lucky_survival_count': getattr(lora, 'lucky_survivals', 0)
        }
        
        if lucky_survived:
            msg = f"\n{'*'*80}\n"
            msg += f"🍀 ŞANSLI KURTULUŞ! (Maç #{self.match_count})\n"
            msg += f"{'*'*80}\n"
            msg += f"  • İsim: {lora.name}\n"
            msg += f"  • Fitness: {lora.get_recent_fitness():.3f} (ölüm eşiğinin altında)\n"
            msg += f"  • Yaş: {self.match_count - lora.birth_match} maç\n"
            msg += f"  • %10 şansı tuttu, hayatta kaldı! 🎲\n"
            msg += f"{'*'*80}\n"
        else:
            age_matches = self.match_count - lora.birth_match
            
            msg = f"\n{'='*80}\n"
            msg += f"💀 LoRA ÖLDÜ (Maç #{self.match_count}){resurrection_tag}\n"
            msg += f"{'='*80}\n"
            msg += f"  • İsim: {lora.name}\n"
            msg += f"  • ID: {lora.id}\n"
            msg += f"  • Yaş: {age_matches} maç\n"
            msg += f"  • Doğum: Maç #{lora.birth_match}\n"
            msg += f"  • Ölüm: Maç #{self.match_count}\n"
            msg += f"  • Final Fitness: {lora.get_recent_fitness():.3f}\n"
            msg += f"  • Generasyon: {lora.generation}\n"
            
            # 🔬 TES SKORLARI!
            msg += f"\n  🔬 TES SKORLARI:\n"
            msg += f"     Total TES: {tes_data['total_tes']:.3f}\n"
            msg += f"     Darwin: {tes_data['darwin']:.2f} | Einstein: {tes_data['einstein']:.2f} | Newton: {tes_data['newton']:.2f}\n"
            msg += f"     Tip: {tes_data['lora_type']}\n"
            
            # ⚡ LIFE ENERGY!
            msg += f"\n  ⚡ YAŞAM ENERJİSİ:\n"
            if life_energy <= 0:
                msg += f"     💀 Enerji tükendi! ({life_energy:.3f})\n"
            else:
                msg += f"     🔋 Son enerji: {life_energy:.3f}\n"
            
            # 🎭 FİZİK ARKETİPİ!
            msg += f"  🎭 Fizik Arketip: {physics_arch}\n"
            msg += f"  🌊 Parçacık Arketip: {particle_arch}\n"
            
            # 🌊 PARÇACIK FİZİĞİ VERİLERİ!
            msg += f"\n  🌊 PARÇACIK FİZİĞİ:\n"
            msg += f"     Sıcaklık (T): {langevin_temp:.4f}\n"
            msg += f"     Sürtünme (ξ): {nose_hoover_xi:.3f}\n"
            msg += f"     Kinetik Enerji: {kinetic_energy:.3f}\n"
            msg += f"     Onsager-Machlup (S_OM): {om_action:.3f}\n"
            msg += f"     Lazarus Λ: {lazarus_lambda:.3f}\n"
            msg += f"     Ghost Potansiyel: {ghost_potential:.3f}\n"
            
            msg += f"\n  • 💀 Ölüm Sebebi: {death_reason_detail or self._get_death_reason_text(reason)}\n"
            msg += f"  • Toplam Maç: {len(lora.match_history)}\n"
            
            # Dirilme geçmişi
            if hasattr(lora, 'resurrection_count') and lora.resurrection_count > 0:
                msg += f"  • ⚡ Dirilme Geçmişi: {lora.resurrection_count} kez dirildi\n"
            
            # Şanslı kurtuluş geçmişi
            if hasattr(lora, 'lucky_survivals') and lora.lucky_survivals > 0:
                msg += f"  • 🍀 Şanslı Kurtuluş: {lora.lucky_survivals} kez ölümden döndü\n"
            
            msg += f"{'='*80}\n"
        
        self._write_log(msg)
        self.all_events.append(event)
        
        # 💀 CANLI EXCEL'E YAZ! (Sadece gerçek ölümlerde!)
        if not lucky_survived:
            try:
                self._log_death_to_excel(event)
            except Exception as e:
                print(f"  ⚠️ Excel'e yazılamadı: {e}")
    
    def log_mating_attempt(self, lora1, lora2, success=True, reason=""):
        """Çiftleşme denemesi logla"""
        event = {
            'match': self.match_count,
            'timestamp': datetime.now().isoformat(),
            'type': 'mating_attempt',
            'lora1_id': lora1.id,
            'lora1_name': lora1.name,
            'lora1_fitness': lora1.get_recent_fitness(),
            'lora2_id': lora2.id,
            'lora2_name': lora2.name,
            'lora2_fitness': lora2.get_recent_fitness(),
            'success': success,
            'reason': reason
        }
        
        if success:
            msg = f"\n💑 ÇİFTLEŞME! (Maç #{self.match_count})\n"
            msg += f"  • Partner 1: {lora1.name} (fitness: {lora1.get_recent_fitness():.3f})\n"
            msg += f"  • Partner 2: {lora2.name} (fitness: {lora2.get_recent_fitness():.3f})\n"
            msg += f"  • Seçim Tipi: {reason}\n"
        else:
            msg = f"\n❌ Çiftleşme başarısız: {reason}\n"
        
        self._write_log(msg)
        self.all_events.append(event)
    
    def log_mutation(self, lora, mutation_type="normal"):
        """Mutasyon logla"""
        event = {
            'match': self.match_count,
            'timestamp': datetime.now().isoformat(),
            'type': 'mutation',
            'lora_id': lora.id,
            'lora_name': lora.name,
            'mutation_type': mutation_type
        }
        
        if mutation_type == "shock":
            msg = f"  ⚡ ŞOK MUTASYON! → {lora.name} (bazı parametreler tamamen yenilendi)\n"
        else:
            msg = f"  🧬 Normal mutasyon → {lora.name}\n"
        
        self._write_log(msg)
        self.all_events.append(event)
    
    def log_match_start(self, match_num, home_team, away_team):
        """Maç başlangıcı"""
        self.match_count = match_num
        
        msg = f"\n\n{'#'*80}\n"
        msg += f"⚽ MAÇ #{match_num}: {home_team} vs {away_team}\n"
        msg += f"{'#'*80}\n"
        
        self._write_log(msg)
    
    def _init_death_report_excel(self):
        """
        Ölüm raporu Excel dosyasını başlat (ilk kez)
        """
        import pandas as pd
        
        # Eğer dosya yoksa, başlık satırıyla oluştur
        if not os.path.exists(self.death_report_file):
            df = pd.DataFrame(columns=[
                'Ölüm Maçı', 'Tarih', 'LoRA İsmi', 'LoRA ID',
                'TES', 'Darwin', 'Einstein', 'Newton', 'Tip',
                'Energy', 'Energy Durum',
                'Sıcaklık(T)', 'Sürtünme(ξ)', 'KE', 'S_OM', 'Lazarus_Λ', 'Ghost_U',  # 🌊🌀🧟👻 YENİ!
                'Yaş (Maç)', 'Yaş (Yıl)', 'Final Fitness', 'Generasyon',
                'Fizik Arketip', 'Parçacık Arketip',  # 🌊 YENİ!
                'Ölüm Sebebi', 'Dirilme Sayısı', 'Şanslı Kurtuluş Sayısı'
            ])
            df.to_excel(self.death_report_file, index=False, engine='openpyxl')
    
    def _log_death_to_excel(self, death_event: dict):
        """
        Ölümü ANINDA Excel'e yaz! (CANLI!)
        
        Args:
            death_event: Ölüm eventi dictionary
        """
        from openpyxl import load_workbook
        
        # Excel'i aç
        wb = load_workbook(self.death_report_file)
        ws = wb.active
        
        # ✅ Yeni satır ekle (PARÇACIK FİZİĞİ! - MAÇ BAZLI!)
        age_in_matches = death_event.get('age_in_matches', 0)
        
        # TES skorları
        tes_scores = death_event.get('tes_scores', {})
        
        # Energy
        life_energy = death_event.get('life_energy', 0.0)
        energy_status = "💀 Tükendi" if life_energy <= 0 else f"🔋 {life_energy:.2f}"
        
        # Parçacık fiziği verileri
        langevin_temp = death_event.get('langevin_temp', 0.01)
        nose_hoover_xi = death_event.get('nose_hoover_xi', 0.0)
        kinetic_energy = death_event.get('kinetic_energy', 0.0)
        om_action = death_event.get('om_action', 0.0)
        lazarus_lambda = death_event.get('lazarus_lambda', 0.5)
        lazarus_lambda = death_event.get('lazarus_lambda', 0.5)
        ghost_potential = death_event.get('ghost_potential', 0.0)
        
        # Yaş hesapla
        age_in_matches = death_event.get('age_in_matches', 0)
        age_in_years = age_in_matches / 34.0  # 1 sezon = 34 maç kabulü

        
        new_row = [
            death_event.get('match', ''),
            death_event.get('timestamp', ''),
            death_event.get('lora_name', ''),
            death_event.get('lora_id', ''),
            tes_scores.get('total_tes', 0.0),
            tes_scores.get('darwin', 0.0),
            tes_scores.get('einstein', 0.0),
            tes_scores.get('newton', 0.0),
            tes_scores.get('lora_type', 'DENGELI⚖️'),
            life_energy,
            energy_status,
            langevin_temp,     # 🌊 Langevin T
            nose_hoover_xi,    # 🌊 Nosé-Hoover ξ
            kinetic_energy,    # 🌊 KE
            om_action,         # 🌀 Onsager-Machlup
            lazarus_lambda,    # 🧟 Lazarus Λ
            ghost_potential,   # 👻 Ghost U
            death_event.get('age_in_matches', 0),
            f"{age_in_years:.1f}",
            death_event.get('final_fitness', 0),
            death_event.get('generation', 0),
            death_event.get('physics_archetype', 'Dengeli⚖️'),
            death_event.get('particle_archetype', 'Dengeli Parçacık ⚖️'),  # 🌊 YENİ!
            death_event.get('death_detail', death_event.get('reason', 'Bilinmiyor')),
            death_event.get('resurrection_count', 0),
            death_event.get('lucky_survival_count', 0)
        ]
        
        ws.append(new_row)
        
        # Kaydet
        wb.save(self.death_report_file)
        wb.close()
    
    def log_era_separator_to_death_report(self, era_type: str = "Normal Run", match_start: int = 0, additional_info: str = ""):
        """
        ÖLÜM RAPORUNA DÖNEM AYIRICI EKLE!
        
        Her yeni başlatma ve diriltmede çağrılır.
        
        Args:
            era_type: "Normal Run", "Resurrection", "Restart", vs.
            match_start: Başlangıç maç numarası
            additional_info: Ek bilgi (örn: "50 LoRA dirildi")
        """
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        # Excel'i aç
        wb = load_workbook(self.death_report_file)
        ws = wb.active
        
        # Boş satır ekle (ayırıcı için)
        ws.append(['', '', '', '', '', '', '', '', '', '', ''])
        
        # DÖNEM BAŞLIK SATIRI
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        separator_text = f"🌅 YENİ DÖNEM: {era_type} | Tarih: {timestamp} | Maç: {match_start}+"
        
        if additional_info:
            separator_text += f" | {additional_info}"
        
        # Başlık satırı (tüm sütunları kaplasın)
        ws.append([separator_text, '', '', '', '', '', '', '', '', '', ''])
        
        # Son eklenen satırı bul ve formatla
        last_row = ws.max_row
        
        # Merge cells (tüm sütunları birleştir)
        ws.merge_cells(f'A{last_row}:K{last_row}')
        
        # Stil uygula
        cell = ws[f'A{last_row}']
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Mavi
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Boş satır ekle (altına)
        ws.append(['', '', '', '', '', '', '', '', '', '', ''])
        
        # Kaydet
        wb.save(self.death_report_file)
        wb.close()
        
        print(f"\n📅 ÖLÜM RAPORUNA DÖNEM AYIRICI EKLENDİ!")
        print(f"   • Tip: {era_type}")
        print(f"   • Tarih: {timestamp}")
        print(f"   • Başlangıç Maç: {match_start}")
    
    def log_resurrection_era(self, resurrected_loras: list, stats: dict):
        """
        Diriltme dönemini kaydet (EXCEL'e yazılacak!)
        
        Args:
            resurrected_loras: Dirilen/spawn edilen LoRA'lar
            stats: Diriltme istatistikleri
        """
        from datetime import datetime
        
        era_info = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'match_num': self.match_count,
            'total_resurrected': len(resurrected_loras),
            'stats': stats,
            'loras': []
        }
        
        # Her LoRA için kayıt
        for lora in resurrected_loras:
            # Tip belirle
            if "Resurrected_" in lora.name:
                lora_type = "Diriltme"
                archetype = "-"
            elif "Balanced_" in lora.name:
                lora_type = "Spawn (Dengeli)"
                # İsimden arketip çıkar
                archetype = lora.name.replace("Balanced_", "").split('_')[0]
            elif "Alien_" in lora.name and lora.name.split('_')[1].isdigit():
                lora_type = "Spawn (Alien)"
                archetype = "Nörotipik Farklılık"
            else:
                lora_type = "Spawn (Uç)"
                # İsimden arketip çıkar
                archetype = lora.name.split('_')[0]
            
            era_info['loras'].append({
                'name': lora.name,
                'id': lora.id,
                'type': lora_type,
                'archetype': archetype,
                'temperament': lora.temperament
            })
        
        self.resurrection_eras.append(era_info)
        
        print(f"📝 Diriltme dönemi kaydedildi: {len(resurrected_loras)} LoRA")
    
    def log_population_snapshot(self, population, additional_stats=None):
        """
        Popülasyon anlık görüntüsü
        Her maçta çağrılır
        """
        if len(population) == 0:
            return
        
        fitnesses = [lora.get_recent_fitness() for lora in population]
        generations = [lora.generation for lora in population]
        ages = [self.match_count - lora.birth_match for lora in population]
        
        snapshot = {
            'match': self.match_count,
            'timestamp': datetime.now().isoformat(),
            'population_size': len(population),
            'avg_fitness': sum(fitnesses) / len(fitnesses),
            'max_fitness': max(fitnesses),
            'min_fitness': min(fitnesses),
            'std_fitness': self._std(fitnesses),
            'avg_generation': sum(generations) / len(generations),
            'max_generation': max(generations),
            'avg_age': sum(ages) / len(ages),
            'max_age': max(ages),
            'oldest_lora': max(population, key=lambda x: self.match_count - x.birth_match).name
        }
        
        if additional_stats:
            snapshot.update(additional_stats)
        
        self.population_history.append(snapshot)
        
        # 🏷️ HER LoRA İÇİN DETAYLI SATIR EKLE (EXCEL İÇİN!)
        for lora in population:
            # Etiketleri al (YENİ METOD!)
            tags = lora.get_status_tags()
            
            # Kişilik tipi
            temp = lora.temperament
            if temp['independence'] > 0.7:
                temp_type = 'Bağımsız'
            elif temp['social_intelligence'] > 0.7:
                temp_type = 'Sosyal Zeki'
            elif temp['herd_tendency'] > 0.6:
                temp_type = 'Sürü'
            elif temp['contrarian_score'] > 0.6:
                temp_type = 'Karşıt'
            else:
                temp_type = 'Dengeli'
            
            # 🏆 İTİBAR HESAPLA!
            from lora_system.reputation_system import reputation_system
            reputation_data = reputation_system.calculate_reputation(
                lora,
                population,
                all_loras_ever=None,
                match_count=self.match_count
            )
            
            # 🎭 DUYGU ARKETİPİ
            emotional_archetype = getattr(lora, 'emotional_archetype', 'Dengeli')
            
            # 🔬 TES SKORLARI!
            from lora_system.tes_scoreboard import tes_scoreboard
            tes_data = tes_scoreboard.calculate_tes_score(lora, population, collective_memory=None)
            
            # 🎭 FİZİK ARKETİPİ (Eski sistem - fluid temperament bazlı)
            from lora_system.physics_based_archetypes import physics_archetypes
            physics_arch = physics_archetypes.determine_archetype_from_physics(lora)
            
            # 🌊 PARÇACIK ARKETİPİ (YENİ! - parçacık fiziği bazlı)
            from lora_system.particle_archetypes import particle_archetypes
            particle_arch_data = particle_archetypes.get_archetype_from_lora(lora)
            particle_arch = particle_arch_data['primary_archetype']
            
            # ⚡ LIFE ENERGY
            life_energy = getattr(lora, 'life_energy', 1.0)
            
            # 🌊 LANGEVIN & NOSÉ-HOOVER!
            langevin_data = {
                'T_eff': getattr(lora, '_langevin_temp', 0.01),
                'xi': getattr(lora, '_nose_hoover_xi', 0.0),
                'KE': getattr(lora, '_kinetic_energy', 0.0)
            }
            
            # 🌀 ONSAGER-MACHLUP!
            om_action = getattr(lora, '_om_action', 0.0)
            
            # 🧟 LAZARUS Λ!
            lazarus_lambda = getattr(lora, '_lazarus_lambda', 0.5)
            
            # 👻 GHOST POTENTIAL!
            ghost_potential = getattr(lora, '_ghost_potential', 0.0)
            
            # Detaylı satır (PARÇACIK FİZİĞİ!)
            row = {
                'Maç': self.match_count,
                'LoRA': lora.name,
                'TES': round(tes_data['total_tes'], 3),
                'Darwin': round(tes_data['darwin'], 2),
                'Einstein': round(tes_data['einstein'], 2),
                'Newton': round(tes_data['newton'], 2),
                'Tip': tes_data['lora_type'],
                'Energy': round(life_energy, 2),
                'Sıcaklık': round(langevin_data['T_eff'], 4),  # 🌊 Langevin T!
                'Sürtünme': round(langevin_data['xi'], 3),     # 🌊 Nosé-Hoover ξ!
                'KE': round(langevin_data['KE'], 3),           # 🌊 Kinetik enerji!
                'S_OM': round(om_action, 3),                   # 🌀 Onsager-Machlup!
                'Lazarus_Λ': round(lazarus_lambda, 3),        # 🧟 Diriltme pot.!
                'Ghost_U': round(ghost_potential, 3),          # 👻 Hayalet pot.!
                'Fitness': round(lora.get_recent_fitness(), 3),
                'Gen': lora.generation,
                'Yaş': self.match_count - lora.birth_match,
                'Uzmanlık': getattr(lora, 'specialization', '-'),
                'Kişilik': temp_type,
                'Duygu Arketip': emotional_archetype,
                'Fizik Arketip': physics_arch,
                'Parçacık Arketip': particle_arch,  # 🌊 YENİ!
                'İtibar': reputation_data['tier'],
                'Etiketler': ' | '.join(tags) if tags else '-',
                
                # 7. Nöral & Genetik Veriler (NEW!)
                'Memory_Size': len(lora.personal_memory_buffer.buffer) if hasattr(lora, 'personal_memory_buffer') else 0,
                'Trauma_Count': len(lora.trauma_history) if hasattr(lora, 'trauma_history') else 0,
                'Children_Count': getattr(lora, 'children_count', 0),
                'Instinct_Source': f"{lora.parents[0]}+{lora.parents[1]}" if len(lora.parents) >= 2 else "None"
            }
            
            self.detailed_lora_history.append(row)
        
        # Her 10 maçta bir detaylı yazdır
        if self.match_count % 10 == 0:
            msg = f"\n{'='*80}\n"
            msg += f"📊 POPÜLASYON DURUMU (Maç #{self.match_count})\n"
            msg += f"{'='*80}\n"
            msg += f"  Popülasyon: {snapshot['population_size']} LoRA\n"
            msg += f"  Fitness:\n"
            msg += f"    • Ortalama: {snapshot['avg_fitness']:.3f}\n"
            msg += f"    • Min/Max: {snapshot['min_fitness']:.3f} / {snapshot['max_fitness']:.3f}\n"
            msg += f"    • Std Dev: {snapshot['std_fitness']:.3f}\n"
            msg += f"  Generasyon:\n"
            msg += f"    • Ortalama: {snapshot['avg_generation']:.1f}\n"
            msg += f"    • Maximum: {snapshot['max_generation']}\n"
            msg += f"  Yaş:\n"
            msg += f"    • Ortalama: {snapshot['avg_age']:.1f} maç\n"
            msg += f"    • En yaşlı: {snapshot['max_age']} maç ({snapshot['oldest_lora']})\n"
            msg += f"{'='*80}\n"
            
            self._write_log(msg)
    
    def log_top_loras(self, population, top_k=5):
        """En iyi LoRA'ları logla"""
        if len(population) == 0:
            return
        
        sorted_pop = sorted(population, key=lambda x: x.get_recent_fitness(), reverse=True)
        
        msg = f"\n🏆 TOP {top_k} LoRA (Maç #{self.match_count}):\n"
        for i, lora in enumerate(sorted_pop[:top_k], 1):
            age = self.match_count - lora.birth_match
            msg += f"  {i}. {lora.name}\n"
            msg += f"     • Fitness: {lora.get_recent_fitness():.3f}\n"
            msg += f"     • Gen: {lora.generation}, Yaş: {age} maç\n"
        
        self._write_log(msg)
    
    def log_prediction(self, base_proba, lora_proba, final_proba, 
                      actual_result=None, correct=None, 
                      predicted_score=None, actual_score=None, score_fitness=None):
        """Tahmin detayları (kazanan + skor)"""
        event = {
            'match': self.match_count,
            'timestamp': datetime.now().isoformat(),
            'type': 'prediction',
            'base_proba': base_proba.tolist() if hasattr(base_proba, 'tolist') else list(base_proba),
            'lora_proba': lora_proba.tolist() if hasattr(lora_proba, 'tolist') else list(lora_proba),
            'final_proba': final_proba.tolist() if hasattr(final_proba, 'tolist') else list(final_proba),
        }
        
        if actual_result:
            event['actual_result'] = actual_result
            event['correct'] = correct
        
        # Skor bilgileri ekle
        if predicted_score:
            event['predicted_score'] = predicted_score
        if actual_score:
            event['actual_score'] = actual_score
        if score_fitness:
            event['score_fitness'] = score_fitness
        
        self.all_events.append(event)
    
    def log_detailed_predictions(self, population, individual_predictions, actual_result, label_encoder):
        """
        Her LoRA'nın tahminini detaylı logla
        
        Args:
            population: LoRA listesi
            individual_predictions: Her LoRA'nın tahmini [(lora, proba), ...]
            actual_result: Gerçek sonuç
            label_encoder: Label encoder
        """
        if len(population) == 0:
            return
        
        actual_idx = list(label_encoder.classes_).index(actual_result)
        
        correct_loras = []
        wrong_loras = []
        
        msg = f"\n🧬 LoRA TAHMİNLERİ ({len(population)} LoRA):\n"
        msg += f"{'='*80}\n"
        
        for lora, proba in individual_predictions:
            pred_idx = proba.argmax()
            pred_class = label_encoder.classes_[pred_idx]
            confidence = proba[pred_idx]
            
            is_correct = (pred_idx == actual_idx)
            status = "✅" if is_correct else "❌"
            
            # Uzmanlık bilgisi (varsa)
            specialization = getattr(lora, 'specialization', None)
            spec_text = f" [{specialization}]" if specialization else ""
            
            msg += f"  {status} {lora.name}{spec_text}: {pred_class} ({confidence*100:.1f}%)"
            msg += f" | Fitness: {lora.get_recent_fitness():.3f}"
            msg += f" | Gen: {lora.generation}"
            msg += f" | Yaş: {self.match_count - lora.birth_match}\n"
            
            if is_correct:
                correct_loras.append((lora, confidence))
            else:
                wrong_loras.append((lora, confidence))
        
        # Özet
        correct_pct = len(correct_loras) / len(population) * 100
        wrong_pct = len(wrong_loras) / len(population) * 100
        
        msg += f"\n📊 ÖZET:\n"
        msg += f"  ✅ Doğru tahmin: {len(correct_loras)}/{len(population)} (%{correct_pct:.1f} bildi)\n"
        msg += f"  ❌ Yanlış tahmin: {len(wrong_loras)}/{len(population)} (%{wrong_pct:.1f} bilemedi)\n"
        msg += f"{'='*80}\n"
        
        self._write_log(msg)
    
    def log_meta_lora_decision(self, attention_weights, population, top_k=5):
        """Meta-LoRA'nın karar sürecini logla"""
        if len(population) == 0 or len(attention_weights) == 0:
            return
        
        # ✅ GÜVENLI KONTROL: attention_weights ve population boyutu uyuşmalı!
        if len(attention_weights) != len(population):
            # Uyuşmuyorsa loglamayı atla
            return
        
        msg = f"\n🧠 META-LoRA KARARI:\n"
        msg += f"{'='*80}\n"
        
        # En yüksek ağırlık alanlar
        sorted_indices = attention_weights.argsort()[::-1]
        
        msg += f"  En yüksek ağırlık alanlar:\n"
        for i in range(min(top_k, len(sorted_indices))):
            idx = sorted_indices[i]
            # ✅ Index kontrolü
            if idx >= len(population):
                continue
            lora = population[idx]
            weight = attention_weights[idx]
            spec = getattr(lora, 'specialization', 'Genel')
            
            msg += f"    {i+1}. {lora.name} ({weight*100:.1f}%)"
            msg += f" - {spec} | Fitness: {lora.get_recent_fitness():.3f}\n"
        
        # En düşük ağırlık
        msg += f"\n  En düşük ağırlık:\n"
        for i in range(min(3, len(sorted_indices))):
            idx = sorted_indices[-(i+1)]
            # ✅ Index kontrolü
            if idx >= len(population):
                continue
            lora = population[idx]
            weight = attention_weights[idx]
            spec = getattr(lora, 'specialization', 'Genel')
            
            msg += f"    • {lora.name} ({weight*100:.1f}%)"
            msg += f" - {spec}\n"
        
        msg += f"{'='*80}\n"
        self._write_log(msg)
    
    def log_nature_graph(self, nature_state, population_size):
        """Doğanın durumunu grafik olarak göster"""
        
        msg = f"\n🌍 DOĞANIN NABZI:\n"
        msg += f"{'='*80}\n"
        
        # Bar grafikler (ASCII)
        health_bar = self._create_bar(nature_state.health, 10, "💚", "░")
        anger_bar = self._create_bar(nature_state.anger, 10, "😡", "░")
        chaos_bar = self._create_bar(nature_state.chaos_index, 10, "🌪️", "░")
        
        msg += f"  Sağlık:  [{health_bar}] {nature_state.health*100:.0f}%\n"
        msg += f"  Öfke:    [{anger_bar}] {nature_state.anger*100:.0f}%\n"
        msg += f"  Kaos:    [{chaos_bar}] {nature_state.chaos_index*100:.0f}%\n"
        msg += f"\n"
        
        # Kara Veba riski (kendi hesaplaması için import gerekli)
        msg += f"  Nüfus: {population_size} LoRA\n"
        
        msg += f"{'='*80}\n"
        self._write_log(msg)
    
    def log_social_bonds(self, lora, population, top_k=3):
        """Bir LoRA'nın sosyal bağlarını logla"""
        
        if not hasattr(lora, 'social_bonds') or len(lora.social_bonds) == 0:
            return
        
        msg = f"\n🔗 SOSYAL BAĞLAR ({lora.name}):\n"
        
        # En güçlü bağlar
        sorted_bonds = sorted(lora.social_bonds.items(), key=lambda x: x[1], reverse=True)
        
        for i, (other_lora_id, strength) in enumerate(sorted_bonds[:top_k]):
            other_lora = next((l for l in population if l.id == other_lora_id), None)
            if other_lora:
                bond_type = self._get_bond_type(strength)
                msg += f"  → {other_lora.name} (çekim: {strength:.2f}) {bond_type}\n"
        
        # Hedefsiz mi?
        if not hasattr(lora, 'main_goal') or lora.main_goal is None:
            goalless_risk = getattr(lora, 'goalless_death_risk', 0.0)
            msg += f"\n  ⚠️ HEDEFSİZ! Sürüklenme riski: {goalless_risk*100:.1f}%\n"
        
        self._write_log(msg)
    
    def log_trauma_history(self, lora, recent_only=True):
        """LoRA'nın travma geçmişini logla"""
        
        if not hasattr(lora, 'trauma_history') or len(lora.trauma_history) == 0:
            return
        
        traumas = lora.trauma_history
        if recent_only:
            # Ciddi travmalar (hem dict hem TraumaEvent)
            traumas = [t for t in traumas if (t.get('severity', 0) if isinstance(t, dict) else t.severity) > 0.3]
        
        if len(traumas) == 0:
            return
        
        msg = f"\n🩹 TRAVMA GEÇMİŞİ ({lora.name}):\n"
        
        for trauma in traumas[-3:]:  # Son 3 travma
            # Trauma hem dict hem TraumaEvent olabilir
            if isinstance(trauma, dict):
                msg += f"  • Maç #{trauma.get('timestamp', trauma.get('match', 0))}: {trauma.get('type', 'unknown')} (şiddet: {trauma.get('severity', 0):.2f})\n"
            else:
                msg += f"  • Maç #{trauma.timestamp}: {trauma.type} (şiddet: {trauma.severity:.2f})\n"
        
        msg += f"\n  Toplam travma: {len(lora.trauma_history)}\n"
        
        self._write_log(msg)
    
    def log_population_graph(self, history, last_n=50):
        """Nüfus grafiği (ASCII)"""
        
        if len(history) < 2:
            return
        
        recent = history[-last_n:] if len(history) > last_n else history
        
        msg = f"\n📈 NÜFUS EVRİMİ (Son {len(recent)} maç):\n"
        msg += f"{'='*80}\n"
        
        sizes = [h['population_size'] for h in recent]
        max_size = max(sizes)
        min_size = min(sizes)
        
        # Y ekseni (5 seviye)
        for level in range(5, 0, -1):
            threshold = min_size + (max_size - min_size) * level / 5
            line = f"{int(threshold):3d} |"
            
            for size in sizes:
                if size >= threshold:
                    line += "●"
                else:
                    line += " "
            
            msg += line + "\n"
        
        msg += f"    |{'_' * len(recent)}\n"
        msg += f"     0{'':>{len(recent)-10}}{len(recent)} (maç)\n"
        
        msg += f"\n  Min: {min_size} | Max: {max_size} | Şu an: {sizes[-1]}\n"
        msg += f"{'='*80}\n"
        
        self._write_log(msg)
    
    def _create_bar(self, value, max_blocks=10, fill_char="█", empty_char="░"):
        """ASCII bar oluştur"""
        filled = int(value * max_blocks)
        empty = max_blocks - filled
        return fill_char * filled + empty_char * empty
    
    def _get_bond_type(self, strength):
        """Bağ gücüne göre emoji"""
        if strength > 0.8:
            return "💚 (çok güçlü)"
        elif strength > 0.6:
            return "💙 (güçlü)"
        elif strength > 0.4:
            return "💛 (orta)"
        elif strength > 0.2:
            return "🧡 (zayıf)"
        elif strength < 0:
            return "💔 (itme/düşmanlık)"
        else:
            return "🤍 (çok zayıf)"
    
    def log_specialization_evolution(self, lora, old_spec, new_spec):
        """Uzmanlık değişimi logla"""
        
        msg = f"\n{'='*80}\n"
        msg += f"🦋 UZMANLIK EVRİMİ! (Maç #{self.match_count})\n"
        msg += f"{'='*80}\n"
        msg += f"  LoRA: {lora.name}\n"
        msg += f"  Generasyon: {lora.generation}\n"
        msg += f"  Yaş: {self.match_count - lora.birth_match} maç\n"
        msg += f"\n"
        msg += f"  Eski Uzmanlık: {old_spec if old_spec else 'Yok'}\n"
        msg += f"  Yeni Uzmanlık: {new_spec}\n"
        msg += f"\n"
        msg += f"  → EVRİM GEÇİRDİ! 🦋\n"
        msg += f"  → Bu LoRA artık farklı pattern'lerde uzmanlaştı!\n"
        msg += f"{'='*80}\n"
        
        self._write_log(msg)
    
    def log_specialization_discovered(self, lora, specialization):
        """İlk uzmanlık keşfi"""
        
        msg = f"\n🎖️ UZMANLIK KEŞFEDİLDİ! (Maç #{self.match_count})\n"
        msg += f"  LoRA: {lora.name}\n"
        msg += f"  Uzmanlık: {specialization}\n"
        msg += f"  → Bu LoRA bu konuda yetenekli!\n\n"
        
        self._write_log(msg)
    
    def log_evolved_loras_summary(self, population):
        """Evrim geçiren tüm LoRA'ları özetle"""
        
        evolved = []
        for lora in population:
            if hasattr(lora, 'specialization_history') and len(lora.specialization_history) > 1:
                evolved.append(lora)
        
        if len(evolved) == 0:
            return
        
        msg = f"\n{'='*80}\n"
        msg += f"🦋 EVRİM GEÇİRENLER ({len(evolved)}/{len(population)} LoRA):\n"
        msg += f"{'='*80}\n"
        
        for lora in evolved:
            msg += f"\n  {lora.name}:\n"
            for i, spec_hist in enumerate(lora.specialization_history):
                duration = "şimdi" if spec_hist.end_match is None else f"{spec_hist.end_match - spec_hist.start_match} maç"
                msg += f"    {i+1}. {spec_hist.specialization} (Maç #{spec_hist.start_match}, süre: {duration})\n"
        
        msg += f"\n{'='*80}\n"
        self._write_log(msg)
    
    def save_all(self):
        """Tüm logları kaydet"""
        # JSON
        with open(self.json_log_file, 'w', encoding='utf-8') as f:
            json.dump({
                'events': self.all_events,
                'population_history': self.population_history,
                'match_count': self.match_count
            }, f, indent=2, ensure_ascii=False)
        
        # Dosya adlarını başta tanımla
        excel_detailed = self.population_csv.replace('.csv', '_DETAYLI.xlsx')
        excel_file = self.population_csv.replace('.csv', '_OZET.xlsx')
        excel_events = self.events_csv.replace('.csv', '.xlsx')
        
        # 🏷️ DETAYLI LoRA EXCEL! (Her LoRA her maçta bir satır + Etiketler + DİRİLTME AYRAÇLARI!)
        if len(self.detailed_lora_history) > 0 or len(self.resurrection_eras) > 0:
            # Excel workbook oluştur
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "LoRA Detaylı"
            
            # Başlık satırı (PARÇACIK FİZİĞİ!)
            headers = [
                'Maç', 'LoRA', 
                'TES', 'Darwin', 'Einstein', 'Newton', 'Tip',
                'Energy', 
                'Sıcaklık(T)', 'Sürtünme(ξ)', 'KE', 'S_OM', 'Lazarus_Λ', 'Ghost_U',  # 🌊🌀🧟👻
                'Fitness', 'Gen', 'Yaş', 'Uzmanlık', 'Kişilik', 
                'Duygu Arketip', 'Fizik Arketip', 'Parçacık Arketip',  # 🌊 YENİ!
                'İtibar', 'Etiketler',
                'Memory_Size', 'Trauma_Count', 'Children_Count', 'Instinct_Source' # 🧠 NEW!
            ]
            ws.append(headers)
            
            # Başlık stili
            for cell in ws[1]:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            current_row = 2
            
            # Diriltme dönemlerini ve normal kayıtları birleştir
            # Önce resurrection_eras'ı işaretle
            resurrection_matches = {era['match_num']: era for era in self.resurrection_eras}
            
            # Maç sırasına göre işle
            last_match = -1
            for record in self.detailed_lora_history:
                match_num = record['Maç']
                
                # Eğer bu maçta diriltme varsa, önce onu yaz!
                if match_num in resurrection_matches and match_num != last_match:
                    era = resurrection_matches[match_num]
                    
                    # AYIRAÇ SATIRI
                    ws.append([])
                    current_row += 1
                    
                    separator_text = f"════ YENİ DÖNEM BAŞLANGICI - {era['timestamp']} ════"
                    ws.merge_cells(f'A{current_row}:J{current_row}')  # H → J (2 sütun eklendi!)
                    cell = ws[f'A{current_row}']
                    cell.value = separator_text
                    cell.font = Font(bold=True, size=14, color="FFFFFF")
                    cell.fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                    current_row += 1
                    
                    # Diriltme detayları
                    ws.append(['DÖNEM', 'LoRA ADI', 'TİP', 'ARKETİP', 'DURUM', '', '', '', '', ''])  # 10 sütun!
                    for cell in ws[current_row]:
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    current_row += 1
                    
                    # Her dirilen LoRA için satır
                    for lora_info in era['loras']:
                        status_emoji = "⚡" if lora_info['type'] == "Diriltme" else "🎭"
                        if "Dengeli" in lora_info['type']:
                            status_emoji = "⚖️"
                        elif "Alien" in lora_info['type']:
                            status_emoji = "👽"
                        
                        ws.append([
                            'YENİ',
                            lora_info['name'],
                            lora_info['type'],
                            lora_info['archetype'],
                            status_emoji,
                            '', '', '', '', ''  # 10 sütun!
                        ])
                        current_row += 1
                    
                    # Ayraç sonrası boş satır
                    ws.append([])
                    current_row += 1
                    
                    last_match = match_num
                
                # Normal maç kaydı (PARÇACIK FİZİĞİ!)
                ws.append([
                    record.get('Maç'),
                    record.get('LoRA'),
                    record.get('TES', 0.50),
                    record.get('Darwin', 0.50),
                    record.get('Einstein', 0.50),
                    record.get('Newton', 0.50),
                    record.get('Tip', 'DENGELI⚖️'),
                    record.get('Energy', 1.0),
                    record.get('Sıcaklık', 0.01),        # 🌊 Langevin T
                    record.get('Sürtünme', 0.0),         # 🌊 Nosé-Hoover ξ
                    record.get('KE', 0.0),               # 🌊 Kinetik enerji
                    record.get('S_OM', 0.0),             # 🌀 Onsager-Machlup
                    record.get('Lazarus_Λ', 0.5),        # 🧟 Lazarus
                    record.get('Ghost_U', 0.0),          # 👻 Ghost potential
                    record.get('Fitness'),
                    record.get('Gen'),
                    record.get('Yaş'),
                    record.get('Uzmanlık'),
                    record.get('Kişilik'),
                    record.get('Duygu Arketip', 'Dengeli'),
                    record.get('Fizik Arketip', 'Dengeli Merkür⚖️'),
                    record.get('Parçacık Arketip', 'Dengeli Parçacık ⚖️'),  # 🌊 YENİ!
                    record.get('İtibar', 'Sıradan'),
                    record.get('Etiketler'),
                    record.get('Memory_Size', 0),
                    record.get('Trauma_Count', 0),
                    record.get('Children_Count', 0),
                    record.get('Instinct_Source', '-')
                ])
                current_row += 1
            
            # Kaydet
            wb.save(excel_detailed)
        
        # ESKİ PD METHODU YOK ARTIK!
        # if len(self.detailed_lora_history) > 0:
        #     df_detailed = pd.DataFrame(self.detailed_lora_history)
        #     df_detailed.to_excel(excel_detailed, index=False, engine='openpyxl')
        
        # Popülasyon Özet EXCEL
        if len(self.population_history) > 0:
            df_pop = pd.DataFrame(self.population_history)
            df_pop.to_excel(excel_file, index=False, engine='openpyxl')
        
        # Events EXCEL
        events_for_excel = [e for e in self.all_events if e['type'] in ['birth', 'death', 'mutation']]
        if len(events_for_excel) > 0:
            df_events = pd.DataFrame(events_for_excel)
            df_events.to_excel(excel_events, index=False, engine='openpyxl')
        
        print(f"\n💾 Loglar kaydedildi:")
        print(f"   • Ana log: {self.main_log_file}")
        print(f"   • JSON: {self.json_log_file}")
        print(f"   • 🏷️ LoRA Detaylı EXCEL: {excel_detailed}")
        print(f"   • Popülasyon Özet: {excel_file}")
        print(f"   • Olaylar EXCEL: {excel_events}")
    
    def _write_log(self, message):
        """Log dosyasına yaz + konsola yazdır"""
        with open(self.main_log_file, 'a', encoding='utf-8') as f:
            f.write(message)
        print(message, end='')
    
    def _get_death_reason_text(self, reason):
        reasons = {
            'low_fitness': 'Düşük fitness (< 0.35)',
            'overpopulation': 'Popülasyon fazlalığı',
            'forced': 'Zorla öldürülme'
        }
        return reasons.get(reason, reason)
    
    def _std(self, values):
        """Standard deviation"""
        if len(values) <= 1:
            return 0.0
        mean = sum(values) / len(values)
        variance = sum((x - mean) ** 2 for x in values) / len(values)
        return variance ** 0.5
    
    def log_specialization_change(self, lora, old_spec, new_spec, match_count):
        """
        Uzmanlık değişimi logla (EVRİM!)
        """
        msg = f"\n{'='*80}\n"
        msg += f"⚡ UZMANLIK EVRİMİ! (Maç #{match_count})\n"
        msg += f"{'='*80}\n"
        msg += f"  LoRA: {lora.name}\n"
        msg += f"  Generasyon: {lora.generation}\n"
        msg += f"  Yaş: {match_count - lora.birth_match} maç\n"
        msg += f"\n"
        
        if old_spec:
            msg += f"  ESKİ UZMANLIK: {old_spec}\n"
        else:
            msg += f"  ESKİ UZMANLIK: Yok (ilk uzmanlık)\n"
        
        msg += f"  YENİ UZMANLIK: {new_spec}\n"
        msg += f"\n"
        msg += f"  🧬 Bu LoRA evrimleşti!\n"
        
        # Pattern başarılarını göster
        if hasattr(lora, 'pattern_stats'):
            msg += f"\n  📊 Pattern Başarıları:\n"
            sorted_patterns = sorted(
                lora.pattern_stats.items(),
                key=lambda x: x[1].rate if x[1].total > 0 else 0,
                reverse=True
            )
            
            for pattern, stats in sorted_patterns[:5]:
                if stats.total > 0:
                    msg += f"    • {pattern}: {stats.rate*100:.1f}% ({stats.correct}/{stats.total})\n"
        
        msg += f"{'='*80}\n"
        
        self._write_log(msg)
        
        # Event'e ekle
        self.all_events.append({
            'match': match_count,
            'type': 'specialization_change',
            'lora_id': lora.id,
            'lora_name': lora.name,
            'old_specialization': old_spec,
            'new_specialization': new_spec,
            'age': match_count - lora.birth_match
        })
    
    def generate_summary_report(self):
        """
        Özet rapor oluştur (TEK DOSYA, APPEND!)
        
        Her çalıştırma sonunda aynı dosyaya AYIRAÇLA ekler!
        """
        if len(self.all_events) == 0:
            return
        
        report_file = os.path.join(self.log_dir, "summary_report.txt")  # ✅ SABİT İSİM!
        
        # İstatistikler
        births = [e for e in self.all_events if e['type'] == 'birth']
        deaths = [e for e in self.all_events if e['type'] == 'death']
        mutations = [e for e in self.all_events if e['type'] == 'mutation']
        
        crossover_births = [b for b in births if b.get('birth_type') == 'crossover']
        spontaneous_births = [b for b in births if b.get('birth_type') == 'spontaneous']
        emergency_births = [b for b in births if b.get('birth_type') == 'emergency']
        
        lucky_survivals = [d for d in deaths if d.get('lucky_survived')]
        
        with open(report_file, 'a', encoding='utf-8') as f:  # ✅ APPEND MODE!
            # AYIRAÇ (her yeni oturumu ayırt et!)
            f.write("\n" + "🔸"*80 + "\n")
            f.write("🔸" + " "*38 + "YENİ OTURUM" + " "*38 + "🔸\n")
            f.write("🔸"*80 + "\n\n")
            
            f.write("=" * 80 + "\n")
            f.write("📊 EVRİM ÖZET RAPORU\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Toplam Maç: {self.match_count}\n")
            f.write(f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            f.write("🐣 DOĞUMLAR:\n")
            f.write(f"  • Toplam: {len(births)}\n")
            f.write(f"  • Çiftleşme: {len(crossover_births)}\n")
            f.write(f"  • Spontane (Alien): {len(spontaneous_births)}\n")
            f.write(f"  • Acil: {len(emergency_births)}\n\n")
            
            f.write("💀 ÖLÜMLER:\n")
            f.write(f"  • Toplam: {len(deaths)}\n")
            f.write(f"  • Şanslı Kurtuluş: {len(lucky_survivals)}\n\n")
            
            f.write("🧬 MUTASYONLAR:\n")
            f.write(f"  • Toplam: {len(mutations)}\n\n")
            
            if len(self.population_history) > 0:
                last = self.population_history[-1]
                first = self.population_history[0]
                
                f.write("📈 POPÜLASYON EVRİMİ:\n")
                f.write(f"  • Başlangıç: {first['population_size']} LoRA\n")
                f.write(f"  • Şu an: {last['population_size']} LoRA\n")
                f.write(f"  • İlk Avg Fitness: {first['avg_fitness']:.3f}\n")
                f.write(f"  • Son Avg Fitness: {last['avg_fitness']:.3f}\n")
                f.write(f"  • İyileşme: {((last['avg_fitness'] - first['avg_fitness']) / first['avg_fitness'] * 100):.1f}%\n")
                f.write(f"  • Max Generation: {last['max_generation']}\n\n")
            
            f.write("=" * 80 + "\n")
        
        print(f"\n📊 Özet rapor oluşturuldu: {report_file}")
    
    def log_hibernation(self, lora, match_idx):
        """
        LoRA hibernation (uyku) logla
        
        KOLONİ MANTIĞI: LoRA ölmez, uyur! Diske kaydedilir, RAM'den çıkar.
        """
        event = {
            'match': match_idx,
            'timestamp': datetime.now().isoformat(),
            'type': 'hibernation',
            'lora_id': lora.id,
            'lora_name': lora.name,
            'age': match_idx - lora.birth_match,
            'fitness': lora.get_recent_fitness() if hasattr(lora, 'get_recent_fitness') else 0.0,
            'generation': lora.generation
        }
        
        msg = f"\n{'~'*80}\n"
        msg += f"😴 LoRA UYUDU (Maç #{match_idx})\n"
        msg += f"{'~'*80}\n"
        msg += f"  • İsim: {lora.name}\n"
        msg += f"  • ID: {lora.id}\n"
        msg += f"  • Fitness: {event['fitness']:.3f}\n"
        msg += f"  • Sebep: Düşük dikkat/performans\n"
        msg += f"  • Durum: Diske kaydedildi, gerektiğinde yüklenecek\n"
        msg += f"  💡 NOT: ÖLMEDI, sadece uyudu! KOLONİ BÜYÜMEYE DEVAM EDER!\n"
        msg += f"{'~'*80}\n"
        
        self.all_events.append(event)
        self._write_log(msg)
    
    def log_miracle_saved(self, lora, match_idx, miracle_id, criteria):
        """
        Mucize LoRA kaydedildi!
        
        🏆 HALL OF FAME
        """
        event = {
            'match': match_idx,
            'timestamp': datetime.now().isoformat(),
            'type': 'miracle',
            'lora_id': lora.id,
            'lora_name': lora.name,
            'miracle_id': miracle_id,
            'miracle_score': criteria['total_points'],
            'fitness': criteria['fitness'],
            'age': criteria['age']
        }
        
        msg = f"\n{'🏆'*80}\n"
        msg += f"🏆 MUCİZE LoRA - HALL OF FAME! (Maç #{match_idx})\n"
        msg += f"{'🏆'*80}\n"
        msg += f"  • İsim: {lora.name}\n"
        msg += f"  • ID: {lora.id}\n"
        msg += f"  • Yaş: {criteria['age']} maç\n"
        msg += f"  • Final Fitness: {criteria['fitness']:.3f}\n"
        msg += f"  • Generasyon: {lora.generation}\n"
        msg += f"  • Uzmanlık: {getattr(lora, 'specialization', 'Genel')}\n"
        msg += f"  • Mucize Puanı: {criteria['total_points']}/100\n"
        msg += f"\n  🌟 SEBEPLER:\n"
        for reason in criteria['reasons']:
            msg += f"      - {reason}\n"
        msg += f"\n  💾 Kaydedildi: {miracle_id}.pt\n"
        msg += f"  📚 Düşünceleri sonsuza kadar korunacak!\n"
        msg += f"{'🏆'*80}\n"
        
        self.all_events.append(event)
        self._write_log(msg)

