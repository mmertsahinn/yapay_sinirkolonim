"""
⚡ YAŞAYAN LoRA'LAR RAPORU (Canlı Excel!)
==========================================

Her 10 maçta güncellenen canlı rapor!

İÇERİK:
- Sadece yaşayan LoRA'lar!
- TES skorları (Darwin, Einstein, Newton)
- Life Energy (Yaşam enerjisi!)
- Fizik arketip
- Grafikler!
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import os
import torch
from typing import List


class LivingLoRAsReporter:
    """
    Yaşayan LoRA'lar için canlı Excel raporu
    """
    
    def __init__(self, report_file: str = "evolution_logs/YASAYAN_LORALAR_CANLI.xlsx"):
        self.report_file = report_file
        
        # İlk dosyayı oluştur
        if not os.path.exists(self.report_file):
            self._create_initial_file()
        
        print(f"⚡ Living LoRAs Reporter başlatıldı: {report_file}")
    
    def _create_initial_file(self):
        """İlk Excel dosyasını oluştur"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Yaşayan LoRAlar"
        
        # Başlık (PARÇACIK FİZİĞİ + HYBRID HİYERARŞİ!)
        headers = [
            'Maç', 'İsim', 'ID', 
            'TES', 'Darwin', 'Einstein', 'Newton', 'Tip', 'Hybrid_Tier',  # 🆕 Hybrid seviyesi!
            'Energy', 'Durum', 
            'T(Sıcaklık)', 'ξ(Sürtünme)', 'KE', 'S_OM', 'Lazarus_Λ', 'Ghost_U',  # 🌊🌀🧟👻
            'Fizik Arketip', 'Fitness', 
            'Yaş', 'Gen', 'Uzmanlık', 'Etiketler'
        ]
        ws.append(headers)
        
        # Başlık stili
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Kaydet
        os.makedirs(os.path.dirname(self.report_file), exist_ok=True)
        wb.save(self.report_file)
        wb.close()
    
    def update_living_loras(self, population: List, match_num: int, tes_scoreboard=None, hibernation_manager=None):
        """
        Yaşayan LoRA'ları Excel'e yaz! (Her 10 maçta)
        
        Args:
            population: Yaşayan (AKTİF) LoRA listesi
            match_num: Maç numarası
            tes_scoreboard: TES scoreboard instance
            hibernation_manager: Hibernation manager (uyuyanları almak için)
        """
        if match_num % 10 != 0:
            return  # Sadece her 10 maçta!
        
        from lora_system.tes_scoreboard import tes_scoreboard as tes_calc
        from lora_system.physics_based_archetypes import physics_archetypes
        import torch
        
        wb = load_workbook(self.report_file)
        ws = wb.active
        
        # 🔥 ESKİ SATIRLARI TEMİZLE! (Sadece başlık kalsın)
        # Excel'de sadece ANLIK yaşayanlar görünsün, biriken log değil!
        ws.delete_rows(2, ws.max_row)  # Satır 2'den itibaren tümünü sil
        
        # 1️⃣ AKTİF LoRA'LARI EKLE
        for lora in population:
            # TES hesapla
            tes_data = tes_calc.calculate_tes_score(lora, population, collective_memory=None)
            
            # Fizik arketip
            physics_arch = physics_archetypes.determine_archetype_from_physics(lora)
            
            # Life energy
            life_energy = getattr(lora, 'life_energy', 1.0)
            
            # Energy durumu
            if life_energy >= 1.5:
                energy_status = "⚡⚡ Yüksek"
            elif life_energy >= 1.0:
                energy_status = "⚡ Normal"
            elif life_energy >= 0.5:
                energy_status = "🔋 Düşük"
            else:
                energy_status = "⚠️ Kritik"
            
            # Etiketler
            tags = []
            if getattr(lora, 'resurrection_count', 0) > 0:
                tags.append(f"⚡x{lora.resurrection_count}")
            if getattr(lora, 'lucky_survivals', 0) > 0:
                tags.append(f"🍀x{lora.lucky_survivals}")
            if hasattr(lora, 'specialization') and lora.specialization:
                tags.append("🎯Uzman")
            
            # 🌊 PARÇACIK FİZİĞİ VERİLERİ
            langevin_temp = getattr(lora, '_langevin_temp', 0.01)
            nose_hoover_xi = getattr(lora, '_nose_hoover_xi', 0.0)
            kinetic_energy = getattr(lora, '_kinetic_energy', 0.0)
            om_action = getattr(lora, '_om_action', 0.0)
            lazarus_lambda = getattr(lora, '_lazarus_lambda', 0.5)
            ghost_potential = getattr(lora, '_ghost_potential', 0.0)
            
            # 🆕 HYBRID TIER BELİRLE!
            hybrid_tier = ""
            if 'PERFECT HYBRID💎💎💎' in tes_data['lora_type']:
                hybrid_tier = "💎 PERFECT"
            elif 'STRONG HYBRID🌟🌟' in tes_data['lora_type']:
                hybrid_tier = "🌟 STRONG"
            elif 'HYBRID🌟' in tes_data['lora_type']:
                hybrid_tier = "⭐ HYBRID"
            elif 'HYBRID' in tes_data['lora_type']:
                hybrid_tier = "🔹 İKİLİ"
            else:
                hybrid_tier = "-"
            
            # Satır ekle (PARÇACIK FİZİĞİ + HYBRID TIER!)
            row = [
                match_num,
                lora.name,
                lora.id[:8],
                round(tes_data['total_tes'], 3),
                round(tes_data['darwin'], 2),
                round(tes_data['einstein'], 2),
                round(tes_data['newton'], 2),
                tes_data['lora_type'],
                hybrid_tier,  # 🆕 Hybrid seviyesi!
                round(life_energy, 2),
                energy_status,
                round(langevin_temp, 4),      # 🌊 T
                round(nose_hoover_xi, 3),     # 🌊 ξ
                round(kinetic_energy, 3),     # 🌊 KE
                round(om_action, 3),          # 🌀 S_OM
                round(lazarus_lambda, 3),     # 🧟 Λ
                round(ghost_potential, 3),    # 👻 U
                physics_arch,
                round(lora.get_recent_fitness(), 3),
                match_num - lora.birth_match,
                lora.generation,
                getattr(lora, 'specialization', '-'),
                ' | '.join(tags) if tags else '-'
            ]
            
            ws.append(row)
        
        # 2️⃣ UYUYAN LoRA'LARI DA EKLE! (😴 UYUYOR durumu)
        hibernated_count = 0
        if hibernation_manager and hasattr(hibernation_manager, 'hibernated_loras'):
            hibernated_ids = list(hibernation_manager.hibernated_loras.keys())
            
            for lora_id, file_path in hibernation_manager.hibernated_loras.items():
                try:
                    # Diskten yükle (metadata için)
                    if os.path.exists(file_path):
                        state = torch.load(file_path, map_location='cpu')
                        metadata = state.get('metadata', {})
                        
                        # Uyuyan LoRA için satır oluştur
                        lora_name = metadata.get('name', f'LoRA_{lora_id[:8]}')
                        birth_match = metadata.get('birth_match', match_num - 100)
                        generation = metadata.get('generation', 0)
                        fitness_history = metadata.get('fitness_history', [])
                        recent_fitness = fitness_history[-1] if fitness_history else 0.5
                        life_energy = metadata.get('life_energy', 1.0)
                        specialization = metadata.get('specialization', '-')
                        
                        # Uyuyan için varsayılan değerler (TES hesaplanmaz, uyuyor!)
                        row = [
                            match_num,
                            lora_name,
                            lora_id[:8],
                            '-',  # TES (uyuyor, hesaplanamaz)
                            '-',  # Darwin
                            '-',  # Einstein
                            '-',  # Newton
                            '😴 UYUYOR',
                            '-',  # Hybrid_Tier
                            round(life_energy, 2),
                            '😴 UYUYOR',  # Durum
                            '-',  # T
                            '-',  # ξ
                            '-',  # KE
                            '-',  # S_OM
                            '-',  # Lazarus_Λ
                            '-',  # Ghost_U
                            '-',  # Fizik Arketip
                            round(recent_fitness, 3),
                            match_num - birth_match,
                            generation,
                            specialization,
                            '😴 UYUYAN'
                        ]
                        
                        ws.append(row)
                        hibernated_count += 1
                except Exception as e:
                    # Hata olursa devam et
                    pass
        
        # Grafik ekle (Her 50 maçta)
        if match_num % 50 == 0 and ws.max_row > 10:
            self._add_energy_chart(ws, wb)
        
        # Kaydet
        wb.save(self.report_file)
        wb.close()
        
        total_living = len(population) + hibernated_count
        print(f"\n⚡ YAŞAYAN LoRA RAPORU GÜNCELLENDİ! (Aktif: {len(population)}, Uyuyan: {hibernated_count}, Toplam: {total_living})")
    
    def _add_energy_chart(self, ws, wb):
        """Energy grafiği ekle"""
        try:
            chart = BarChart()
            chart.title = "Yaşam Enerjisi Dağılımı"
            chart.x_axis.title = "LoRA"
            chart.y_axis.title = "Energy"
            
            # Tüm satırları al (Artık sadece o anki yaşayanlar var!)
            max_row = ws.max_row
            
            # Data reference (Energy sütunu = 9)
            # Başlık hariç tüm satırlar
            data = Reference(ws, min_col=9, min_row=2, max_row=max_row)
            categories = Reference(ws, min_col=2, min_row=2, max_row=max_row)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(categories)
            
            # Grafik ekle
            ws.add_chart(chart, f"R{max_row+2}")
        except Exception as e:
            pass  # Grafik eklenemezse devam et


# Global instance
living_reporter = LivingLoRAsReporter()

